#!/usr/bin/env python3
import json, os, io, smtplib, base64, urllib.request, urllib.error, hashlib, secrets, traceback, hmac, time
from datetime import datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from http.server import ThreadingHTTPServer, BaseHTTPRequestHandler
from urllib.parse import urlparse, parse_qs
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR  = os.path.dirname(os.path.abspath(__file__))
PTPL_SOL  = os.path.join(BASE_DIR, 'plantillas', 'Solicitud_plantilla.xlsx')
PTPL_REP  = os.path.join(BASE_DIR, 'plantillas', 'Reposicion_plantilla.xlsx')
LOGO_PATH = os.path.join(BASE_DIR, 'plantillas', 'logo.png')

GMAIL_USER   = os.environ.get('GMAIL_USER', '')
GMAIL_PASS   = os.environ.get('GMAIL_PASS', '')
GITHUB_TOKEN = os.environ.get('GITHUB_TOKEN', '')
GITHUB_OWNER = os.environ.get('GITHUB_OWNER', 'DIRECCIONGESTIONDEFONDOS')
GITHUB_REPO  = os.environ.get('GITHUB_REPO',  'VIATICOS-SINALOA')
DATA_FILE    = 'data.json'
NET_TIMEOUT  = int(os.environ.get('NET_TIMEOUT', '20'))

ADMIN_EMAILS = [
    e.strip().lower()
    for e in os.environ.get(
        'ADMIN_EMAILS',
        'abraham.navarro@sinaloa.gob.mx,direcciongestiondefondos@gmail.com'
    ).split(',')
    if e.strip()
]
DEFAULT_ADMIN_PASS = os.environ.get('ADMIN_PASSWORD', 'DGFADMIN')
RESET_ADMIN_PASSWORD = os.environ.get('RESET_ADMIN_PASSWORD', 'false').strip().lower() in ('1', 'true', 'yes', 'si', 'sí')

# Clave para firmar sesiones sin guardarlas en memoria.
# Configúrala en Render como SESSION_SECRET.
SESSION_SECRET = os.environ.get('SESSION_SECRET') or hashlib.sha256(
    (GITHUB_TOKEN or 'viaticos-sinaloa').encode('utf-8')
).hexdigest()

# ── SESIONES SIN MEMORIA / TOKEN FIRMADO ──────────────────────────────────────
def hash_pass(p):
    return hashlib.sha256(p.encode()).hexdigest()

def _b64url_encode(raw):
    return base64.urlsafe_b64encode(raw).decode('utf-8').rstrip('=')

def _b64url_decode(txt):
    padding = '=' * (-len(txt) % 4)
    return base64.urlsafe_b64decode((txt + padding).encode('utf-8'))

def create_session(email, role):
    payload = {
        'email': str(email).strip().lower(),
        'role': role or 'user',
        'exp': int(time.time()) + (8 * 60 * 60)
    }
    payload_b64 = _b64url_encode(json.dumps(payload, separators=(',', ':')).encode('utf-8'))
    sig = hmac.new(SESSION_SECRET.encode('utf-8'), payload_b64.encode('utf-8'), hashlib.sha256).digest()
    return f"{payload_b64}.{_b64url_encode(sig)}"

def get_session(token):
    if not token or '.' not in token:
        return None
    try:
        payload_b64, sig_b64 = token.split('.', 1)
        expected_sig = hmac.new(SESSION_SECRET.encode('utf-8'), payload_b64.encode('utf-8'), hashlib.sha256).digest()
        actual_sig = _b64url_decode(sig_b64)
        if not hmac.compare_digest(expected_sig, actual_sig):
            return None
        payload = json.loads(_b64url_decode(payload_b64).decode('utf-8'))
        if int(payload.get('exp', 0)) < int(time.time()):
            return None
        return {'email': payload.get('email', '').lower(), 'role': payload.get('role', 'user')}
    except Exception:
        return None

def clean_sessions():
    # Compatibilidad: ya no hay sesiones en memoria que limpiar.
    return

# ── DÍAS HÁBILES ──────────────────────────────────────────────────────────────
FESTIVOS = {(1,1),(2,5),(3,21),(9,16),(11,2),(11,20),(12,25)}

def es_habil(d):
    if d.weekday() >= 5: return False
    return (d.month, d.day) not in FESTIVOS

def restar_habiles(fecha, n):
    d = fecha
    while n > 0:
        d -= timedelta(days=1)
        if es_habil(d): n -= 1
    return d

# ── NÚMERO A LETRAS ───────────────────────────────────────────────────────────
def num_letras(n):
    u = ['','UN','DOS','TRES','CUATRO','CINCO','SEIS','SIETE','OCHO','NUEVE','DIEZ',
         'ONCE','DOCE','TRECE','CATORCE','QUINCE','DIECISÉIS','DIECISIETE','DIECIOCHO','DIECINUEVE']
    d = ['','','VEINTE','TREINTA','CUARENTA','CINCUENTA','SESENTA','SETENTA','OCHENTA','NOVENTA']
    c = ['','CIENTO','DOSCIENTOS','TRESCIENTOS','CUATROCIENTOS','QUINIENTOS',
         'SEISCIENTOS','SETECIENTOS','OCHOCIENTOS','NOVECIENTOS']
    if n == 0: return 'CERO'
    r = ''
    if n >= 1000:
        m = int(n//1000); r += ('MIL ' if m==1 else num_letras(m)+' MIL '); n = int(n%1000)
    if n >= 100:
        r += ('CIEN ' if n==100 else c[int(n//100)]+' '); n = int(n%100)
    if n >= 20: r += d[int(n//10)] + (' Y '+u[int(n%10)] if n%10>0 else '') + ' '
    elif n > 0: r += u[int(n)] + ' '
    return r.strip()

def monto_letras(m):
    e = int(m); cts = round((m-e)*100)
    return num_letras(e) + (f' CON {cts:02d}/100' if cts>0 else '') + ' M.N.'

# ── RFC EN CELDAS ──────────────────────────────────────────────────────────────
def rfc_sol(rfc):
    ch = list(rfc.upper().replace(' ',''))
    m = {}
    for i,col in enumerate(['C','D','E','F','G','H','I','J','K']):
        m[col+'21'] = ch[i] if i<len(ch) else ''
    m['P21'] = ch[9]  if len(ch)>9  else ''
    m['Q21'] = ''.join(ch[10:]) if len(ch)>10 else ''
    return m

def rfc_rep(rfc):
    ch = list(rfc.upper().replace(' ',''))
    m = {}
    for i,col in enumerate(['B','C','D','E','F','G','H','I','J']):
        m[col+'21'] = ch[i] if i<len(ch) else ''
    m['O21'] = ch[9]  if len(ch)>9  else ''
    m['P21'] = ''.join(ch[10:]) if len(ch)>10 else ''
    return m

# ── GITHUB STORAGE ────────────────────────────────────────────────────────────
def gh_headers():
    return {
        'Authorization': f'token {GITHUB_TOKEN}',
        'Accept': 'application/vnd.github.v3+json',
        'Content-Type': 'application/json',
        'User-Agent': 'viaticos-app'
    }

def asegurar_admins(data):
    """
    Normaliza data.json para que la app siempre reciba listas válidas.
    También asegura administradores iniciales y el programa base EQUIPA SINALOA.
    Corrige registros viejos que hayan quedado como {"id": "...", "datos": {...}}.
    """
    if not isinstance(data, dict):
        data = {}

    changed = False

    # Asegurar estructura base
    defaults = {
        'perfiles': [],
        'programas': [],
        'folios': {},
        'historial': [],
        'usuarios': []
    }
    for k, v in defaults.items():
        if k not in data or data.get(k) is None:
            data[k] = v
            changed = True

    if not isinstance(data.get('perfiles'), list):
        data['perfiles'] = []
        changed = True
    if not isinstance(data.get('programas'), list):
        data['programas'] = []
        changed = True
    if not isinstance(data.get('usuarios'), list):
        data['usuarios'] = []
        changed = True
    if not isinstance(data.get('historial'), list):
        data['historial'] = []
        changed = True
    if not isinstance(data.get('folios'), dict):
        data['folios'] = {}
        changed = True

    # Normalizar perfiles
    perfiles_norm = []
    for idx, p in enumerate(data.get('perfiles', [])):
        if not isinstance(p, dict):
            changed = True
            continue

        # Corrige registros mal guardados como {id:'p...', datos:{nombre...}}
        if isinstance(p.get('datos'), dict):
            merged = dict(p.get('datos'))
            merged['id'] = p.get('id') or merged.get('id') or f'p_{idx+1}'
            p = merged
            changed = True

        p.setdefault('id', f'p_{idx+1}')
        p.setdefault('nombre', '')
        p.setdefault('rfc', '')
        p.setdefault('email', '')
        p.setdefault('cargo', '')
        p.setdefault('rango', '')
        p.setdefault('area', '')
        p.setdefault('auth1_nombre', '')
        p.setdefault('auth1_cargo', '')
        p.setdefault('auth2_nombre', '')
        p.setdefault('auth2_cargo', '')
        perfiles_norm.append(p)

    if perfiles_norm != data.get('perfiles', []):
        data['perfiles'] = perfiles_norm
        changed = True

    # Normalizar programas
    programas_norm = []
    for idx, p in enumerate(data.get('programas', [])):
        if not isinstance(p, dict):
            changed = True
            continue

        p.setdefault('id', f'prog_{idx+1}')
        p.setdefault('nombre', '')
        if not isinstance(p.get('motivos'), list):
            p['motivos'] = []
            changed = True

        # Quitar programas totalmente vacíos
        if not str(p.get('nombre', '')).strip():
            changed = True
            continue

        programas_norm.append(p)

    data['programas'] = programas_norm

    # Asegurar EQUIPA SINALOA como programa base editable
    equipa_base = {
        'id': 'prog_equipa_sinaloa',
        'nombre': 'EQUIPA SINALOA',
        'motivos': [
            'Entrega de equipamiento productivo a personas beneficiarias',
            'Supervisión y seguimiento a beneficiarios del programa',
            'Validación de solicitudes y expedientes del programa',
            'Levantamiento de información para integración de padrones',
            'Reunión de coordinación operativa del programa'
        ]
    }

    idx_equipa = next(
        (
            i for i, p in enumerate(data['programas'])
            if str(p.get('id', '')).strip() == 'prog_equipa_sinaloa'
            or str(p.get('nombre', '')).strip().lower() == 'equipa sinaloa'
        ),
        None
    )

    if idx_equipa is None:
        data['programas'].insert(0, equipa_base)
        changed = True
    else:
        p = data['programas'][idx_equipa]
        if not p.get('id'):
            p['id'] = 'prog_equipa_sinaloa'
            changed = True
        if not p.get('nombre'):
            p['nombre'] = 'EQUIPA SINALOA'
            changed = True
        if not isinstance(p.get('motivos'), list) or not p.get('motivos'):
            p['motivos'] = equipa_base['motivos']
            changed = True
        # mover al inicio
        if idx_equipa != 0:
            equipa = data['programas'].pop(idx_equipa)
            data['programas'].insert(0, equipa)
            changed = True

    # Asegurar administradores iniciales
    nombres_default = {
        'abraham.navarro@sinaloa.gob.mx': 'Abraham Navarro',
        'direcciongestiondefondos@gmail.com': 'Dirección Gestión de Fondos'
    }

    for correo in ADMIN_EMAILS:
        usuario = next(
            (
                u for u in data['usuarios']
                if isinstance(u, dict) and str(u.get('email', '')).strip().lower() == correo
            ),
            None
        )
        if not usuario:
            data['usuarios'].append({
                'email': correo,
                'pass_hash': hash_pass(DEFAULT_ADMIN_PASS),
                'role': 'admin',
                'nombre': nombres_default.get(correo, correo)
            })
            changed = True
        else:
            if usuario.get('role') != 'admin':
                usuario['role'] = 'admin'
                changed = True
            if RESET_ADMIN_PASSWORD or not usuario.get('pass_hash'):
                nuevo_hash = hash_pass(DEFAULT_ADMIN_PASS)
                if usuario.get('pass_hash') != nuevo_hash:
                    usuario['pass_hash'] = nuevo_hash
                    changed = True
            if not usuario.get('nombre'):
                usuario['nombre'] = nombres_default.get(correo, correo)
                changed = True

    return data, changed

def gh_get_data():
    url = f'https://api.github.com/repos/{GITHUB_OWNER}/{GITHUB_REPO}/contents/{DATA_FILE}'
    req = urllib.request.Request(url, headers=gh_headers())
    try:
        with urllib.request.urlopen(req, timeout=NET_TIMEOUT) as r:
            resp = json.loads(r.read())
            content = base64.b64decode(resp['content']).decode('utf-8')
            data = json.loads(content)
            data['_sha'] = resp['sha']
            data, changed = asegurar_admins(data)
            if changed:
                saved = gh_save_data(data.copy())
                if isinstance(saved, dict) and saved.get('content', {}).get('sha'):
                    data['_sha'] = saved['content']['sha']
            return data
    except urllib.error.HTTPError as e:
        if e.code == 404:
            data = {
                'perfiles': [],
                'programas': [],
                'folios': {},
                'historial': [],
                'usuarios': [],
                '_sha': None
            }
            data, _ = asegurar_admins(data)
            return data
        raise

def gh_save_data(data):
    sha = data.pop('_sha', None)
    content = base64.b64encode(json.dumps(data, ensure_ascii=False, indent=2).encode()).decode()
    payload = {'message': 'update data', 'content': content}
    if sha: payload['sha'] = sha
    url = f'https://api.github.com/repos/{GITHUB_OWNER}/{GITHUB_REPO}/contents/{DATA_FILE}'
    req = urllib.request.Request(url, data=json.dumps(payload).encode(),
                                  headers=gh_headers(), method='PUT')
    with urllib.request.urlopen(req, timeout=NET_TIMEOUT) as r:
        return json.loads(r.read())

def siguiente_folio():
    year = str(datetime.now().year)
    data = gh_get_data()
    folios = data.get('folios', {})
    n = folios.get(year, 0) + 1
    folios[year] = n
    data['folios'] = folios
    gh_save_data(data)
    return f'SE-SSGFF-{n:04d}/{year}', n

# ── LOGO EN EXCEL ──────────────────────────────────────────────────────────────
def agregar_logo(ws):
    if not os.path.exists(LOGO_PATH): return
    try:
        img = XLImage(LOGO_PATH)
        img.width = 320; img.height = 72
        img.anchor = 'B1'
        ws.add_image(img)
    except Exception as e:
        print(f"Logo error: {e}")

# ── GENERAR EXCEL ──────────────────────────────────────────────────────────────
def generar_solicitud(d):
    wb  = load_workbook(PTPL_SOL)
    ws  = wb['MAZATLAN']
    agregar_logo(ws)
    total  = float(d.get('total', 0))
    nombre = d.get('nombre','').upper()
    ws['B11'] = d.get('titulo','SOLICITUD DE OFICIO DE COMISIÓN')
    ws['Q12'] = d.get('folio','')
    ws['Q14'] = d.get('fecha_exp','')
    ws['E18'] = d.get('programa','')
    for cel, val in rfc_sol(d.get('rfc','')).items(): ws[cel] = val
    ws['R21'] = nombre
    ws['C22'] = d.get('cargo','').upper()
    ws['B25'] = d.get('destino','').upper()
    ws['R25'] = d.get('fecha_ini','')
    ws['B27'] = d.get('motivo','').upper()
    ws['R27'] = d.get('fecha_fin','')
    letras = monto_letras(total)
    ws['B29'] = (f"RECIBÍ DE LA SECRETARÍA DE ECONOMÍA LA CANTIDAD DE ${total:,.2f} "
                 f"SON: ({letras}) POR CONCEPTO DE  VIATICOS  PARA CUMPLIR LAS COMISIONES ARRIBA INDICADAS.")
    ws['R29'] = d.get('alimentos')   or ''
    ws['R30'] = d.get('gasolina')    or ''
    ws['R31'] = d.get('peaje')       or ''
    ws['R32'] = d.get('hospedaje')   or ''
    ws['B33'] = f"__________________________________________________                    {nombre}"
    ws['B41'] = d.get('auth1_nombre','').upper()
    ws['P41'] = d.get('auth2_nombre','').upper()
    ws['B42'] = d.get('auth1_cargo','').upper()
    ws['P42'] = d.get('auth2_cargo','').upper()
    for sn in list(wb.sheetnames):
        if sn != 'MAZATLAN': del wb[sn]
    ws.title = 'Solicitud'
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()

def generar_reposicion(d):
    wb  = load_workbook(PTPL_REP)
    ws  = wb['reposición']
    agregar_logo(ws)
    total  = float(d.get('total', 0))
    nombre = d.get('nombre','').upper()
    ws['P12'] = d.get('folio','')
    ws['P14'] = d.get('fecha_exp','')
    ws['D18'] = d.get('programa','')
    for cel, val in rfc_rep(d.get('rfc','')).items(): ws[cel] = val
    ws['Q21'] = nombre
    ws['B22'] = d.get('cargo','').upper()
    ws['A25'] = d.get('destino','').upper()
    ws['Q25'] = d.get('fecha_ini','')
    ws['A27'] = d.get('motivo','').upper()
    ws['Q27'] = d.get('fecha_fin','')
    letras = monto_letras(total)
    ws['A29'] = (f"RECIBÍ DE LA SECRETARIA DE ECONOMÍA LA CANTIDAD DE ${total:,.2f} "
                 f"(SON: {letras}), POR CONCEPTO DE  VIÁTICOS,  PARA CUMPLIR LAS COMISIONES ARRIBA INDICADAS.")
    ws['Q29'] = d.get('alimentos')   or ''
    ws['Q30'] = d.get('hospedaje')   or ''
    ws['Q31'] = d.get('combustible') or ''
    ws['Q32'] = d.get('peaje')       or ''
    ws['Q33'] = d.get('transporte')  or ''
    ws['A33'] = nombre
    ws['A43'] = d.get('auth1_nombre','').upper()
    ws['O43'] = d.get('auth2_nombre','').upper()
    ws['A44'] = d.get('auth1_cargo','').upper()
    ws['O44'] = d.get('auth2_cargo','').upper()
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()

# ── HISTORIAL EXCEL ───────────────────────────────────────────────────────────
def generar_historial_excel(historial):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Historial de Solicitudes'
    headers = ['Folio','Fecha Exp.','Nombre','RFC','Cargo','Destino','Fecha Ini','Fecha Fin','Programa','Total','Tipo','Correo Enviado']
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='0F2543')
        cell.alignment = Alignment(horizontal='center')
    for row_i, r in enumerate(historial, 2):
        ws.cell(row=row_i, column=1,  value=r.get('folio',''))
        ws.cell(row=row_i, column=2,  value=r.get('fecha_exp',''))
        ws.cell(row=row_i, column=3,  value=r.get('nombre',''))
        ws.cell(row=row_i, column=4,  value=r.get('rfc',''))
        ws.cell(row=row_i, column=5,  value=r.get('cargo',''))
        ws.cell(row=row_i, column=6,  value=r.get('destino',''))
        ws.cell(row=row_i, column=7,  value=r.get('fecha_ini',''))
        ws.cell(row=row_i, column=8,  value=r.get('fecha_fin',''))
        ws.cell(row=row_i, column=9,  value=r.get('programa',''))
        ws.cell(row=row_i, column=10, value=float(r.get('total',0)))
        ws.cell(row=row_i, column=11, value=r.get('tipo',''))
        ws.cell(row=row_i, column=12, value='Sí' if r.get('email_ok') else 'No')
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()

# ── CORREO ────────────────────────────────────────────────────────────────────
def html_correo(d, total, tipo_doc):
    def fmt(n): return f"${float(n):,.2f}" if n else None
    filas = [
        ("Alimentos",   fmt(d.get('alimentos'))),
        ("Gasolina",    fmt(d.get('gasolina') or d.get('combustible'))),
        ("Peaje",       fmt(d.get('peaje'))),
        ("Hospedaje",   fmt(d.get('hospedaje'))),
    ]
    conceptos_html = ''.join(
        f'<tr><td style="padding:8px 12px;border-bottom:1px solid #eee;color:#555;">{l}</td>'
        f'<td style="padding:8px 12px;border-bottom:1px solid #eee;text-align:right;font-weight:600;">{v}</td></tr>'
        for l,v in filas if v
    )
    nombre_corto = d.get('nombre','').title().split()[0]
    return f"""<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;background:#f4f6fa;font-family:Arial,sans-serif;">
<table width="100%" cellpadding="0" cellspacing="0" style="padding:32px 0;">
<tr><td align="center">
<table width="600" cellpadding="0" cellspacing="0" style="background:#fff;border-radius:12px;overflow:hidden;box-shadow:0 4px 24px rgba(0,0,0,.08);">
  <tr><td style="background:#0f2543;padding:28px 36px;">
    <div style="font-size:11px;letter-spacing:.15em;color:#c8a84b;text-transform:uppercase;margin-bottom:6px;">Secretaría de Economía · Sinaloa</div>
    <div style="font-size:20px;font-weight:700;color:#c8a84b;">{tipo_doc}</div>
    <div style="font-size:13px;color:rgba(255,255,255,.5);margin-top:4px;">Folio: {d.get('folio','—')}</div>
  </td></tr>
  <tr><td style="padding:32px 36px;">
    <p style="color:#555;font-size:14px;margin:0 0 20px;">Estimado(a) <strong>{nombre_corto}</strong>,</p>
    <p style="color:#555;font-size:14px;margin:0 0 24px;">Se adjunta tu solicitud de viáticos con el siguiente detalle:</p>
    <div style="background:#f0f5ff;border-radius:8px;padding:16px 20px;margin-bottom:20px;">
      <div style="font-size:11px;font-weight:700;letter-spacing:.1em;text-transform:uppercase;color:#c8a84b;margin-bottom:12px;">Datos de la comisión</div>
      <table width="100%" cellpadding="0" cellspacing="0" style="font-size:13px;color:#333;">
        <tr><td style="padding:4px 0;color:#888;width:140px;">Destino</td><td style="font-weight:600;">{d.get('destino','—').title()}</td></tr>
        <tr><td style="padding:4px 0;color:#888;">Fecha inicial</td><td style="font-weight:600;">{d.get('fecha_ini','—')}</td></tr>
        <tr><td style="padding:4px 0;color:#888;">Fecha final</td><td style="font-weight:600;">{d.get('fecha_fin','—')}</td></tr>
        <tr><td style="padding:4px 0;color:#888;">Programa</td><td style="font-weight:600;">{d.get('programa','—')}</td></tr>
        <tr><td style="padding:4px 0;color:#888;">Motivo</td><td style="font-weight:600;">{d.get('motivo','—').capitalize()}</td></tr>
      </table>
    </div>
    <div style="font-size:11px;font-weight:700;letter-spacing:.1em;text-transform:uppercase;color:#c8a84b;margin-bottom:10px;">Desglose de viáticos</div>
    <table width="100%" cellpadding="0" cellspacing="0" style="font-size:14px;border:1px solid #eee;border-radius:8px;overflow:hidden;">
      {conceptos_html}
      <tr style="background:#0f2543;">
        <td style="padding:12px;color:#c8a84b;font-weight:700;text-transform:uppercase;font-size:13px;">TOTAL</td>
        <td style="padding:12px;color:#c8a84b;font-weight:700;font-size:18px;text-align:right;">${total:,.2f}</td>
      </tr>
    </table>
    <p style="color:#aaa;font-size:11px;margin:24px 0 0;line-height:1.6;">Correo generado automáticamente · Sistema de Viáticos · Secretaría de Economía Sinaloa</p>
  </td></tr>
  <tr><td style="background:#f9f9f9;padding:14px 36px;border-top:1px solid #eee;text-align:center;">
    <div style="font-size:11px;color:#aaa;">Secretaría de Economía · Gobierno del Estado de Sinaloa · {datetime.now().year}</div>
  </td></tr>
</table></td></tr></table></body></html>"""

def enviar_correo(dest, asunto, html, xlsx_bytes, filename):
    msg = MIMEMultipart('mixed')
    msg['From']    = f"Viáticos SE Sinaloa <{GMAIL_USER}>"
    msg['To']      = dest
    msg['Subject'] = asunto
    alt = MIMEMultipart('alternative')
    alt.attach(MIMEText(html, 'html', 'utf-8'))
    msg.attach(alt)
    adj = MIMEBase('application','vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    adj.set_payload(xlsx_bytes)
    encoders.encode_base64(adj)
    adj.add_header('Content-Disposition','attachment',filename=filename)
    msg.attach(adj)
    with smtplib.SMTP_SSL('smtp.gmail.com', 465, timeout=NET_TIMEOUT) as s:
        s.login(GMAIL_USER, GMAIL_PASS)
        s.sendmail(GMAIL_USER, dest, msg.as_string())

# ── HTTP HANDLER ──────────────────────────────────────────────────────────────
HTML_PATH = os.path.join(BASE_DIR, 'index.html')

def serve_static(path):
    files = {
        '/manifest.json': ('application/json', os.path.join(BASE_DIR, 'manifest.json')),
        '/sw.js': ('application/javascript', os.path.join(BASE_DIR, 'sw.js')),
        '/icon-192.png': ('image/png', os.path.join(BASE_DIR, 'icon-192.png')),
        '/icon-512.png': ('image/png', os.path.join(BASE_DIR, 'icon-512.png')),
        '/logo.png': ('image/png', os.path.join(BASE_DIR, 'logo.png')),
    }
    return files.get(path)

def get_token_from_req(handler):
    cookie = handler.headers.get('Cookie', '')
    for part in cookie.split(';'):
        part = part.strip()
        if part.startswith('session='):
            return part[8:]
    return None

class Handler(BaseHTTPRequestHandler):
    def log_message(self, fmt, *args): pass

    def cors(self):
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'POST,GET,OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type,Authorization')

    def send_json(self, obj, code=200):
        try:
            body = json.dumps(obj, ensure_ascii=False).encode()
            self.send_response(code)
            self.send_header('Content-Type', 'application/json; charset=utf-8')
            self.send_header('Content-Length', len(body))
            self.cors()
            self.end_headers()
            self.wfile.write(body)
        except (BrokenPipeError, ConnectionResetError) as e:
            print(f"Cliente cerró la conexión antes de recibir respuesta: {e}", flush=True)
        except Exception:
            print("Error dentro de send_json:", flush=True)
            print(traceback.format_exc(), flush=True)

    def read_body(self):
        n = int(self.headers.get('Content-Length', 0))
        return json.loads(self.rfile.read(n)) if n else {}

    def get_auth(self):
        # Try Authorization header first
        auth = self.headers.get('Authorization', '')
        if auth.startswith('Bearer '):
            return get_session(auth[7:])
        # Try cookie
        return get_session(get_token_from_req(self))

    def do_OPTIONS(self):
        self.send_response(200); self.cors(); self.end_headers()

    def do_GET(self):
        path = urlparse(self.path).path

        # Static files
        static = serve_static(path)
        if static:
            mime, fpath = static
            if os.path.exists(fpath):
                with open(fpath, 'rb') as f: content = f.read()
                self.send_response(200)
                self.send_header('Content-Type', mime)
                self.send_header('Content-Length', len(content))
                self.cors(); self.end_headers(); self.wfile.write(content)
            else:
                self.send_response(404); self.end_headers()
            return

        if path in ('/', '/index.html'):
            with open(HTML_PATH, 'rb') as f: content = f.read()
            self.send_response(200)
            self.send_header('Content-Type', 'text/html; charset=utf-8')
            self.send_header('Content-Length', len(content))
            self.cors(); self.end_headers(); self.wfile.write(content)
            return

        # API: get data (requires auth)
        if path == '/api/data':
            sess = self.get_auth()
            if not sess:
                self.send_json({'error': 'no_auth'}, 401); return
            try:
                data = gh_get_data()
                result = {
                    'perfiles': data.get('perfiles', []),
                    'programas': data.get('programas', []),
                    'role': sess['role'],
                    'email': sess['email'],
                }
                if sess['role'] == 'admin':
                    result['usuarios'] = data.get('usuarios', [])
                    result['historial'] = data.get('historial', [])
                self.send_json(result)
            except Exception as e:
                print('ERROR DETALLADO:', flush=True)
                print(traceback.format_exc(), flush=True)
                self.send_json({'error': str(e)}, 500)
            return

        # API: delete historial entry (admin only)
        if path == '/api/historial/delete':
            sess = self.get_auth()
            if not sess or sess['role'] != 'admin':
                self.send_json({'error': 'forbidden'}, 403); return
            folio = self.read_body().get('folio','')
            if not folio:
                self.send_json({'ok': False, 'error': 'folio requerido'}); return
            try:
                data = gh_get_data()
                sha  = data.get('_sha')
                antes = len(data.get('historial', []))
                data['historial'] = [r for r in data.get('historial', []) if r.get('folio') != folio]
                despues = len(data['historial'])
                data['_sha'] = sha
                gh_save_data(data)
                self.send_json({'ok': True, 'eliminados': antes - despues})
            except Exception as e:
                print('ERROR DETALLADO:', flush=True)
                print(traceback.format_exc(), flush=True)
                self.send_json({'ok': False, 'error': str(e)}, 500)
            return

        # API: delete ALL historial (admin only)
        if path == '/api/historial/clear':
            sess = self.get_auth()
            if not sess or sess['role'] != 'admin':
                self.send_json({'error': 'forbidden'}, 403); return
            try:
                data = gh_get_data()
                sha  = data.get('_sha')
                data['historial'] = []
                data['folios'] = {}
                data['_sha'] = sha
                data, _ = asegurar_admins(data)
                gh_save_data(data)
                self.send_json({'ok': True})
            except Exception as e:
                print('ERROR DETALLADO:', flush=True)
                print(traceback.format_exc(), flush=True)
                self.send_json({'ok': False, 'error': str(e)}, 500)
            return

        # API: download historial (admin only)
        if path == '/api/historial':
            sess = self.get_auth()
            if not sess or sess['role'] != 'admin':
                self.send_json({'error': 'forbidden'}, 403); return
            try:
                data = gh_get_data()
                xlsx = generar_historial_excel(data.get('historial', []))
                fecha = datetime.now().strftime('%Y%m%d')
                fname = f'Historial_Viaticos_{fecha}.xlsx'
                self.send_response(200)
                self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                self.send_header('Content-Disposition', f'attachment; filename="{fname}"')
                self.send_header('Content-Length', len(xlsx))
                self.cors(); self.end_headers(); self.wfile.write(xlsx)
            except Exception as e:
                print('ERROR DETALLADO:', flush=True)
                print(traceback.format_exc(), flush=True)
                self.send_json({'error': str(e)}, 500)
            return

        self.send_response(404); self.end_headers()

    def do_POST(self):
        path = urlparse(self.path).path
        d    = self.read_body()

        # ── LOGIN ──────────────────────────────────────────────────────────────
        if path == '/api/login':
            email = d.get('email','').lower().strip()
            pwd   = d.get('password','')
            try:
                data     = gh_get_data()
                usuarios = data.get('usuarios', [])
                usuario  = next((u for u in usuarios if u['email'].lower() == email), None)
                if not usuario:
                    self.send_json({'ok': False, 'error': 'Correo no registrado. Contacta al administrador.'}); return
                if not usuario.get('pass_hash'):
                    self.send_json({'ok': False, 'first_time': True, 'error': 'first_time'}); return
                if usuario.get('pass_hash') != hash_pass(pwd):
                    self.send_json({'ok': False, 'error': 'Contrasena incorrecta'}); return
                role  = usuario.get('role', 'user')
                token = create_session(email, role)
                self.send_json({'ok': True, 'token': token, 'role': role, 'nombre': usuario.get('nombre','')})
            except Exception as e:
                print('ERROR DETALLADO:', flush=True)
                print(traceback.format_exc(), flush=True)
                self.send_json({'ok': False, 'error': str(e)}, 500)
            return

        # ── PRIMERA VEZ: CREAR CONTRASENA ─────────────────────────────────────
        if path == '/api/set-password':
            email = d.get('email','').lower().strip()
            pwd   = d.get('password','')
            if len(pwd) < 6:
                self.send_json({'ok': False, 'error': 'Minimo 6 caracteres'}); return
            try:
                data     = gh_get_data()
                sha      = data.get('_sha')
                usuarios = data.get('usuarios', [])
                usuario  = next((u for u in usuarios if u['email'].lower() == email), None)
                if not usuario:
                    self.send_json({'ok': False, 'error': 'Usuario no encontrado'}); return
                if usuario.get('pass_hash'):
                    self.send_json({'ok': False, 'error': 'Este usuario ya tiene contrasena'}); return
                usuario['pass_hash'] = hash_pass(pwd)
                data['_sha'] = sha
                gh_save_data(data)
                role  = usuario.get('role', 'user')
                token = create_session(email, role)
                self.send_json({'ok': True, 'token': token, 'role': role, 'nombre': usuario.get('nombre','')})
            except Exception as e:
                print('ERROR DETALLADO:', flush=True)
                print(traceback.format_exc(), flush=True)
                self.send_json({'ok': False, 'error': str(e)}, 500)
            return

        # ── CAMBIAR CONTRASEÑA ─────────────────────────────────────────────────
        if path == '/api/cambiar-pass':
            sess = self.get_auth()
            if not sess:
                self.send_json({'ok': False, 'error': 'no_auth'}, 401); return
            old_pwd = d.get('old_password','')
            new_pwd = d.get('new_password','')
            try:
                data = gh_get_data()
                sha  = data.get('_sha')
                usuarios = data.get('usuarios', [])
                usuario  = next((u for u in usuarios if u['email'].lower() == sess['email']), None)
                if not usuario or usuario.get('pass_hash') != hash_pass(old_pwd):
                    self.send_json({'ok': False, 'error': 'Contraseña actual incorrecta'}); return
                usuario['pass_hash'] = hash_pass(new_pwd)
                data['_sha'] = sha
                data, _ = asegurar_admins(data)
                gh_save_data(data)
                self.send_json({'ok': True})
            except Exception as e:
                print('ERROR DETALLADO:', flush=True)
                print(traceback.format_exc(), flush=True)
                self.send_json({'ok': False, 'error': str(e)}, 500)
            return

        # ── ADMIN: GESTIÓN DE USUARIOS ─────────────────────────────────────────
        if path == '/api/usuarios/save':
            sess = self.get_auth()
            if not sess or sess['role'] != 'admin':
                self.send_json({'ok': False, 'error': 'forbidden'}, 403); return
            try:
                data = gh_get_data()
                sha  = data.get('_sha')
                data['usuarios'] = d.get('usuarios', data.get('usuarios', []))
                data['_sha'] = sha
                data, _ = asegurar_admins(data)
                gh_save_data(data)
                self.send_json({'ok': True})
            except Exception as e:
                print('ERROR DETALLADO:', flush=True)
                print(traceback.format_exc(), flush=True)
                self.send_json({'ok': False, 'error': str(e)}, 500)
            return

        # Todas las demás rutas requieren auth
        sess = self.get_auth()
        if not sess:
            self.send_json({'error': 'no_auth'}, 401); return

        # ── GUARDAR DATOS ──────────────────────────────────────────────────────
        if path == '/api/save':
            try:
                data = gh_get_data()
                sha  = data.get('_sha')
                # Usuarios: solo admin puede modificar
                if sess['role'] == 'admin':
                    for key in ('perfiles','programas','folios'):
                        if key in d: data[key] = d[key]
                else:
                    # Usuario normal: puede agregar perfiles y programas pero NO editar
                    if 'perfiles_add' in d:
                        data.setdefault('perfiles', []).append(d['perfiles_add'])
                    if 'programas' in d:
                        data['programas'] = d['programas']
                data['_sha'] = sha
                data, _ = asegurar_admins(data)
                gh_save_data(data)
                self.send_json({'ok': True})
            except Exception as e:
                print('ERROR DETALLADO:', flush=True)
                print(traceback.format_exc(), flush=True)
                self.send_json({'ok': False, 'error': str(e)}, 500)
            return

        # ── GENERAR DOCUMENTO ──────────────────────────────────────────────────
        if path in ('/generar/solicitud', '/generar/reposicion'):
            print(f"Generando documento: {path} por {sess.get('email','')}", flush=True)
            try:
                es_rep = path == '/generar/reposicion'
                folio, _ = siguiente_folio()
                d['folio'] = folio

                if d.get('fecha_ini'):
                    try:
                        fi = datetime.strptime(d['fecha_ini'], '%d/%m/%Y')
                        fe = restar_habiles(fi, 3)
                        d['fecha_exp'] = fe.strftime('%d/%m/%Y')
                    except: pass

                print('Generando Excel...', flush=True)
                xlsx   = generar_reposicion(d) if es_rep else generar_solicitud(d)
                print('Excel generado.', flush=True)
                tipo   = d.get('titulo','SOLICITUD DE OFICIO DE COMISIÓN')
                total  = float(d.get('total', 0))
                nom    = d.get('nombre','').replace(' ','_')[:18]
                dst    = d.get('destino','').replace(' ','_')[:12]
                prefix = 'Reposicion' if es_rep else 'Solicitud'
                fname  = f"{prefix}_{nom}_{dst}_{folio.replace('/','-')}.xlsx"

                print('Procesando correo...', flush=True)
                # Correo
                email_dest = d.get('email_destinatario','').strip()
                email_ok = False; email_err = ''
                if email_dest and GMAIL_USER and GMAIL_PASS:
                    try:
                        asunto = f"Viáticos — {tipo} | {d.get('destino','').title()} | {d.get('fecha_ini','')}"
                        enviar_correo(email_dest, asunto, html_correo(d, total, tipo), xlsx, fname)
                        email_ok = True
                    except Exception as e:
                        email_err = str(e)
                        print('ERROR EN CORREO:', flush=True)
                        print(traceback.format_exc(), flush=True)

                print('Guardando historial...', flush=True)
                # Guardar en historial
                try:
                    data = gh_get_data()
                    sha  = data.get('_sha')
                    data.setdefault('historial', []).append({
                        'folio': folio,
                        'fecha_exp': d.get('fecha_exp',''),
                        'nombre': d.get('nombre',''),
                        'rfc': d.get('rfc',''),
                        'cargo': d.get('cargo',''),
                        'destino': d.get('destino',''),
                        'fecha_ini': d.get('fecha_ini',''),
                        'fecha_fin': d.get('fecha_fin',''),
                        'programa': d.get('programa',''),
                        'total': total,
                        'tipo': 'Reposición' if es_rep else 'Oficio',
                        'email_ok': email_ok,
                        'generado_por': sess['email'],
                        'timestamp': datetime.now().isoformat()
                    })
                    data['_sha'] = sha
                    gh_save_data(data)
                except Exception as e:
                    print(f"Error guardando historial: {e}")

                print('Enviando respuesta al navegador...', flush=True)
                self.send_json({
                    'filename': fname, 'folio': folio,
                    'fecha_exp': d.get('fecha_exp',''),
                    'email_ok': email_ok, 'email_err': email_err,
                    'xlsx_b64': base64.b64encode(xlsx).decode()
                })
            except Exception as e:
                print('ERROR DETALLADO:', flush=True)
                print(traceback.format_exc(), flush=True)
                self.send_json({'error': str(e)}, 500)
            return

        self.send_response(404); self.end_headers()


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 8080))
    print(f"Servidor iniciado en puerto {port}")
    ThreadingHTTPServer(('0.0.0.0', port), Handler).serve_forever()
